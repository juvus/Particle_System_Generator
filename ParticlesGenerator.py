#==========================================================================================
# Project:        Particles Generator Tool
# Description:    Program for generation of particles system with desired
#                 distribution of the main parameters
# Author:         Dmitry Safonov (Dmitry.Safonov@lut.fi)
# Organization:   Lappeenranta University of Technology (LUT)
#                 Solid/Liquid Separation Research Group            
# Last edited:    26.12.2020
#==========================================================================================

import sys
import os
import numpy as np
from PyQt5.QtWidgets import (QMainWindow, QApplication, QDesktopWidget, QAction, 
                             QLabel, QPushButton, QMessageBox, QFileDialog, QCheckBox)
from PyQt5.QtGui import QIcon, QFont, QPixmap
from PyQt5.QtCore import Qt, QThreadPool, QTimer
from openpyxl import Workbook
from time import localtime, strftime, time, sleep
import random as rnd
import math as m
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from subprocess import Popen, CREATE_NEW_CONSOLE

# Program modules:
from Modules.About import About
from Modules.AdvancedQSpinBox import AdvancedQSpinBox
from Modules.AdvancedQLineEdit import AdvancedQLineEdit
from Modules.AdvancedQProgressBar import AdvancedQProgressBar
from Modules.Particle import Particle
from Modules.PSOAlg_dll import PSOAlg_dll
from Modules.ImageLabelGenerator import ImageLabelGenerator
from Modules.PSOSettingsWindow import PSOSettingsWindow
from Modules.PSearchSettingsWindow import PSearchSettingsWindow
from Modules.Worker import Worker
from Modules.EditValidateFcn import edit_str_to_value

# Main window class     
class ParticlesGenerator(QMainWindow):  
    """Class of a main window of the particle generator tool"""
    def __init__(self):
        """Constructor of the class"""
        super().__init__()
        rnd.seed()
        self.nDim = 12  # Number of particle dimensions
        self.particle = Particle()  # Construct particle for determination of its parameters
        # Distributions
        # x values (All):
        self.CEDiam_chLower = np.zeros(100)  # Lower channel numbers of CE diameter distribution
        self.CEDiam_chCentre = np.zeros(100)  # Center channel numbers of CE diameter distribution
        self.CEDiam_chUpper = np.zeros(100)  # Upper channel numbers of CE diameter distribution
        self.cirConEl_chLower = np.zeros(100)  # Lower channel numbers of circ., convex. and elong. distributions
        self.cirConEl_chCentre = np.zeros(100)  # Center channel numbers of circ., convex. and elong. distributions
        self.cirConEl_chUpper = np.zeros(100)  # Upper channel numbers of circ., convex. and elong. distributions
        # y values (CEDiameter)
        self.init_CEDiam_distr_cum = np.zeros(100)  # Initial CE diameter cumulative distribution values
        self.init_CEDiam_distr_diff = None  # Initial CE diameter differential distribution values
        self.norm_CEDiam_distr_diff = None  # Normalized CE diameter differential distribution values
        self.count_CEDiam_distr_diff = None  # Additional temp CE diameter differential distribution values
        self.gen_CEDiam_distr_cum = None  # Generated CE diameter cumulative distribution values
        self.gen_CEDiam_distr_diff = None  # Generated CE diameter differential distribution values
        # y values (Circularity)
        self.init_circ_distr_cum = np.zeros(100)  # Initial circularity cumulative distribution values
        self.init_circ_distr_diff = None  # Initial circularity differential distribution values
        self.norm_circ_distr_diff = None  # Normalized circularity differential distribution values
        self.count_circ_distr_diff = None  # Additional temp circularity differential distribution values
        self.gen_circ_distr_cum = None  # Generated circularity cumulative distribution values
        self.gen_circ_distr_diff = None  # Generated circularity differential distribution values
        # y values (Convexity)
        self.init_convex_distr_cum = np.zeros(100)  # Initial convexity cumulative distribution values
        self.init_convex_distr_diff = None  # Initial convexity differential distribution values
        self.norm_convex_distr_diff = None  # Normalized convexity differential distribution values
        self.count_convex_distr_diff = None  # Additional temp convexity differential distribution values
        self.gen_convex_distr_cum = None  # Generated convexity cumulative distribution values
        self.gen_convex_distr_diff = None  # Generated convexity differential distribution values
        # y values (Elongation)
        self.init_elong_distr_cum = np.zeros(100)  # Initial elongation cumulative distribution values
        self.init_elong_distr_diff = None  # Initial elongation differential distribution values
        self.norm_elong_distr_diff = None  # Normalized elongation differential distribution values
        self.count_elong_distr_diff = None  # Additional temp elongation differential distribution values
        self.gen_elong_distr_cum = None  # Generated elongation cumulative distribution values
        self.gen_elong_distr_diff = None  # Generated elongation differential distribution values
        # y values (Solidity)
        self.count_solid_distr_diff = None  # Additional temp solidity differential distribution values
        self.gen_solid_distr_cum = None  # Generated solidity cumulative distribution values
        self.gen_solid_distr_diff = None  # Generated solidity differential distribution values
        # Differential distributions boundaries
        self.CEDiam_leftBndChannel = 0
        self.CEDiam_rightBndChannel = 0
        self.circ_leftBndChannel = 0
        self.circ_rightBndChannel = 0
        self.convex_leftBndChannel = 0
        self.convex_rightBndChannel = 0
        self.elong_leftBndChannel = 0
        self.elong_rightBndChannel = 0
        # Target particle parameters (for the search):
        self.target_CEDiameter = None  # Target CE diameter of the particle, [um]
        self.target_circularity = None  # Target particle circularity, [-]
        self.target_convexity = None  # Target particle convexity, [-]
        self.target_elongation = None  # Target particle elongation, [-]
        # Generated particle parameters (after the search):
        self.gen_dims = None  # Dimensions of the generated particle
        self.gen_CEDiameter = None  # Generated CE diameter of the particle, [um]
        self.gen_circularity = None  # Generated particle circularity, [-]
        self.gen_convexity = None  # Generated particle convexity, [-]
        self.gen_elongation = None  # Generated particle elongation, [-]
        self.gen_solidity = None  # Generated particle solidity, [-]           
        # Sum of area in um2
        self.sumAreaUm2 = None  #  Sum of all generated particles areas (in um2)
        # Particles system properties
        self.onlySpherical = False  # Flag to generate only spherical particles
        self.picturesNum = None  # Number of generated pictures with particles
        self.partPerPicture = None  # Number of particles per picture
        self.particlesNum = None  # Total number of generated particles
        # Search parameters:
        self.stopGeneration = False  # Flag to stop the generation
        self.lamp = False  # Flag showing the work of the green lamp
        self.iterLimit = None  # PSO iteration limit
        self.precisionLimit = None  # PSO precision limit
        self.genStartTime = None  # Generation start time
        self.elapsedTime = None  # Elapsed time of the generation in seconds
        self.timeToFinish = None  # Seconds to finish the particles generation
        self.generatedParts = None  # Total amount of generated particles
        self.percentComplete = None  # Percent to complete the search
        self.useParallelSearch = None  # Flag to use the parallel search
        self.deleteError = None  # Flag of error to delete the files in "data" folder
        self.numThreads = None  # Number of searching threads
        # PSO optimization algorithm hyper parameters:
        self.psoAlg_dll = PSOAlg_dll()  # Instance of the PSO algorithm class (C code from dll)
        self.PSO_nVar = None  # Number of unknown (decision) variables (equal to nDim)
        self.PSO_varMin = None  # Lower bound of decision variables
        self.PSO_varMax = None  # Upper bound of decision variables
        self.PSO_nPop = None  # Population size (swarm size)
        self.PSO_w = None  # Inertia coefficient
        self.PSO_wDamp = None  # Damping ratio of inertia coefficient
        self.PSO_c1 = None  # Personal acceleration coefficient
        self.PSO_c2 = None  # Social acceleration coefficient
        self.PSO_a = None  # Additional randomization of a-th particle in swarm
        self.PSO_b = None  # Additional randomization of all particles every b-th iteration
        # Generated particle image and properties:
        self.showGeneratedPlots = False  # Flag to show the generated distributions plots
        self.fileName = None  # Filename
        self.threadpool = QThreadPool()
        self.init_ui()  # Initialize the user interface elements

    def init_ui(self):
        """Method for the initialization of the UI"""
        self.setFixedSize(574, 535)  # Window size
        self.center_window()  # Center the window on desktop
        self.setWindowIcon(QIcon('Resources/icon.png'))
        self.setWindowTitle('Particles generator tool')  # Window title
        # Menu and submenu creation 
        mainMenu = self.menuBar() 
        distrMenu = mainMenu.addMenu('Distributions')
        infoMenu = mainMenu.addMenu('Info')
        # Menu action - Distributions -> Load file...
        self.loadDistrAct = QAction(QIcon('Resources/load.png'), 'Open file...', self)
        self.loadDistrAct.setShortcut('Ctrl+O')
        self.loadDistrAct.setFont(QFont('Arial', 11))
        self.loadDistrAct.triggered.connect(self.load_distr_data)
        # Menu action - Distributions -> Show plots
        self.showPlotsAct = QAction(QIcon('Resources/plot.png'), 'Show plots...', self)
        self.showPlotsAct.setShortcut('Ctrl+P')
        self.showPlotsAct.setFont(QFont('Arial', 11))
        self.showPlotsAct.triggered.connect(self.show_distr_plots)
        self.showPlotsAct.setDisabled(True)     
        # Menu action: Info -> Help 
        helpAct = QAction(QIcon('Resources/help.png'), 'Help', self)
        helpAct.setFont(QFont('Arial', 11))
        helpAct.setShortcut('Ctrl+H')
        helpAct.triggered.connect(self.show_help)
        # Menu action Info -> About
        aboutAct = QAction(QIcon('Resources/about.png'), 'About', self)
        aboutAct.setFont(QFont('Arial', 11))
        aboutAct.setShortcut('Ctrl+A')
        aboutAct.triggered.connect(self.show_about)
        
        # Add actions to menu: Distributions -> ...
        distrMenu.addAction(self.loadDistrAct)
        distrMenu.addAction(self.showPlotsAct)
        # Add actions to menu: Info -> ...
        infoMenu.addAction(helpAct)
        infoMenu.addAction(aboutAct)
        
        # Label (Particle system properties)
        self.lbl_PSProps = QLabel('Particles system properties:', self)
        self.lbl_PSProps.setGeometry(20, 30, 220, 20)
        self.lbl_PSProps.setFont(QFont('Arial', 11, weight=QFont.Bold))           
        # Label (Search algorithm)
        self.lbl_searchAlg = QLabel('Search algorithm:', self)
        self.lbl_searchAlg.setGeometry(20, 194, 220, 20)
        self.lbl_searchAlg.setFont(QFont('Arial', 11, weight=QFont.Bold))     
        # Label (Generated particle image)
        self.lbl_generatedImg = QLabel('Generated particle image:', self)
        self.lbl_generatedImg.setGeometry(310, 30, 220, 20)
        self.lbl_generatedImg.setFont(QFont('Arial', 11, weight=QFont.Bold))              
        # Label (Generation information)
        self.lbl_generationInfo = QLabel('Generation information:', self)
        self.lbl_generationInfo.setGeometry(20, 332, 220, 20)
        self.lbl_generationInfo.setFont(QFont('Arial', 11, weight=QFont.Bold))

        """Particle system properties section"""
        self.chb_onlySpherical = QCheckBox('Only spherical particles', self)
        self.chb_onlySpherical.setGeometry(20, 53, 200, 21)
        self.chb_onlySpherical.setFont(QFont('Arial', 11))
        self.chb_onlySpherical.clicked.connect(self.chb_onlySpherical_clicked)

        # Block with axes number changer
        self.lbl_axesNum = QLabel('Axes number:', self)
        self.lbl_axesNum.setAlignment(Qt.AlignLeft)
        self.lbl_axesNum.setGeometry(20, 79, 95, 21)
        self.lbl_axesNum.setFont(QFont('Arial', 11))
        
        self.spb_axesNum = AdvancedQSpinBox(self)
        self.spb_axesNum.setAlignment(Qt.AlignRight)
        self.spb_axesNum.setGeometry(163, 79, 65, 21)
        self.spb_axesNum.setFont(QFont('Arial', 11))
        self.spb_axesNum.setRange(3, 32)
        self.spb_axesNum.valueChanged.connect(self.val_changed_spb_axesNum)
        
        self.btn_resetAxesNum = QPushButton(self)
        self.btn_resetAxesNum.setGeometry(244, 79, 21, 21)
        self.btn_resetAxesNum.setIcon(QIcon(QPixmap('./Resources/arrow.png')))
        self.btn_resetAxesNum.clicked.connect(self.reset_axesNum)
        
        # Block with number of pictures
        self.lbl_picturesNum = QLabel('Number of pictures:', self)
        self.lbl_picturesNum.setAlignment(Qt.AlignLeft)
        self.lbl_picturesNum.setGeometry(20, 102, 150, 21)
        self.lbl_picturesNum.setFont(QFont('Arial', 11))
        
        self.spb_picturesNum = AdvancedQSpinBox(self)
        self.spb_picturesNum.setAlignment(Qt.AlignRight)
        self.spb_picturesNum.setGeometry(163, 102, 65, 21)
        self.spb_picturesNum.setFont(QFont('Arial', 11))
        self.spb_picturesNum.setRange(1, 9000)
        self.spb_picturesNum.valueChanged.connect(self.update_edt_particlesNum)
        
        self.btn_resetPicturesNum = QPushButton(self)
        self.btn_resetPicturesNum.setGeometry(244, 102, 21, 21)
        self.btn_resetPicturesNum.setIcon(QIcon(QPixmap('./Resources/arrow.png')))
        self.btn_resetPicturesNum.clicked.connect(self.reset_picturesNum)
        
        # Block with particles per picture
        self.lbl_partPerPicture = QLabel('Particles in picture:', self)
        self.lbl_partPerPicture.setAlignment(Qt.AlignLeft)
        self.lbl_partPerPicture.setGeometry(20, 125, 150, 21)
        self.lbl_partPerPicture.setFont(QFont('Arial', 11))

        self.spb_partPerPicture = AdvancedQSpinBox(self)
        self.spb_partPerPicture.setAlignment(Qt.AlignRight)
        self.spb_partPerPicture.setGeometry(163, 125, 65, 21)
        self.spb_partPerPicture.setFont(QFont('Arial', 11))
        self.spb_partPerPicture.setRange(1, 9000)
        self.spb_partPerPicture.valueChanged.connect(self.update_edt_particlesNum)

        self.btn_resetPartPerPicture = QPushButton(self)
        self.btn_resetPartPerPicture.setGeometry(244, 125, 21, 21)
        self.btn_resetPartPerPicture.setIcon(QIcon(QPixmap('./Resources/arrow.png')))
        self.btn_resetPartPerPicture.clicked.connect(self.reset_partPerPicture)

        # Block with particle total number
        self.lbl_particlesNum = QLabel('Particles number:', self)
        self.lbl_particlesNum.setAlignment(Qt.AlignLeft)
        self.lbl_particlesNum.setGeometry(20, 148, 150, 21)
        self.lbl_particlesNum.setFont(QFont('Arial', 11))
        
        self.edt_particlesNum = AdvancedQLineEdit(self)
        self.edt_particlesNum.setAlignment(Qt.AlignRight)
        self.edt_particlesNum.setGeometry(163, 148, 101, 21)
        self.edt_particlesNum.setFont(QFont('Arial', 11))
        self.edt_particlesNum.set_readOnly(True)
         
        """Search algorithm properties section"""
        # Label of search algorithm
        self.lbl_algorithmName = QLabel('Particle swarm optimization', self)
        self.lbl_algorithmName.setAlignment(Qt.AlignLeft)
        self.lbl_algorithmName.setGeometry(20, 217, 200, 21)
        self.lbl_algorithmName.setFont(QFont('Arial', 11))
        
        # Buttons to open the searching algorithm settings
        self.btn_PSOAlgSettings = QPushButton(self)
        self.btn_PSOAlgSettings.setGeometry(244, 217, 21, 21)
        self.btn_PSOAlgSettings.setIcon(QIcon(QPixmap('./Resources/settings.png')))
        self.btn_PSOAlgSettings.clicked.connect(self.open_PSOAlg_settings)
        
        # Block with iterations limit:
        self.lbl_iterLimit = QLabel('Set iteration limit:', self)
        self.lbl_iterLimit.setAlignment(Qt.AlignLeft)
        self.lbl_iterLimit.setGeometry(20, 240, 150, 21)
        self.lbl_iterLimit.setFont(QFont('Arial', 11))

        self.spb_iterLimit = AdvancedQSpinBox(self)
        self.spb_iterLimit.setAlignment(Qt.AlignRight)
        self.spb_iterLimit.setGeometry(163, 240, 65, 21)
        self.spb_iterLimit.setFont(QFont('Arial', 11))
        self.spb_iterLimit.setRange(1, 9000)
        self.spb_iterLimit.valueChanged.connect(self.val_changed_spb_iterLimit)

        self.btn_resetIterLimit = QPushButton(self)
        self.btn_resetIterLimit.setGeometry(244, 240, 21, 21)
        self.btn_resetIterLimit.setIcon(QIcon(QPixmap('./Resources/arrow.png')))
        self.btn_resetIterLimit.clicked.connect(self.reset_iterLimit)
        
        # Block with precision limit
        self.lbl_precisionLimit = QLabel('Set precision limit:', self)
        self.lbl_precisionLimit.setAlignment(Qt.AlignLeft)
        self.lbl_precisionLimit.setGeometry(20, 263, 150, 21)
        self.lbl_precisionLimit.setFont(QFont('Arial', 11))
        
        self.edt_precisionLimit = AdvancedQLineEdit(self)
        self.edt_precisionLimit.setAlignment(Qt.AlignRight)
        self.edt_precisionLimit.setGeometry(163, 263, 65, 21)
        self.edt_precisionLimit.setFont(QFont('Arial', 11))
        self.edt_precisionLimit.editingFinished.connect(self.update_precisionLimit)
        
        self.btn_resetPrecisionLimit = QPushButton(self)
        self.btn_resetPrecisionLimit.setGeometry(244, 263, 21, 21)
        self.btn_resetPrecisionLimit.setIcon(QIcon(QPixmap('./Resources/arrow.png')))
        self.btn_resetPrecisionLimit.clicked.connect(self.reset_precisionLimit)
        
        # Block with parallel search settings
        self.chb_useParallelSearch = QCheckBox('Use parallel search', self)
        self.chb_useParallelSearch.setGeometry(20, 286, 200, 21)
        self.chb_useParallelSearch.setFont(QFont('Arial', 11))
        self.chb_useParallelSearch.clicked.connect(self.chb_useParallelSearch_clicked)
        
        self.btn_ParallelSearchSettings = QPushButton(self)
        self.btn_ParallelSearchSettings.setGeometry(244, 286, 21, 21)
        self.btn_ParallelSearchSettings.setIcon(QIcon(QPixmap('./Resources/settings.png')))
        self.btn_ParallelSearchSettings.clicked.connect(self.open_ParallelSearch_settings)
        
        """Generation information section"""
        # Block with generation start time
        self.lbl_StartDateTime = QLabel('Started date/time:', self)
        self.lbl_StartDateTime.setAlignment(Qt.AlignLeft)
        self.lbl_StartDateTime.setGeometry(20, 355, 150, 21)
        self.lbl_StartDateTime.setFont(QFont('Arial', 11))
        
        self.edt_startDateTime = AdvancedQLineEdit(self)
        self.edt_startDateTime.setAlignment(Qt.AlignRight)
        self.edt_startDateTime.setGeometry(163, 355, 101, 21)
        self.edt_startDateTime.setFont(QFont('Arial', 11))
        self.edt_startDateTime.set_readOnly(True)
        
        # Block with elapsed time
        self.lbl_elapsedTime = QLabel('Elapsed time:', self)
        self.lbl_elapsedTime.setAlignment(Qt.AlignLeft)
        self.lbl_elapsedTime.setGeometry(20, 378, 150, 21)
        self.lbl_elapsedTime.setFont(QFont('Arial', 11))
        
        self.edt_elapsedTime = AdvancedQLineEdit(self)
        self.edt_elapsedTime.setAlignment(Qt.AlignRight)
        self.edt_elapsedTime.setGeometry(163, 378, 101, 21)
        self.edt_elapsedTime.setFont(QFont('Arial', 11))
        self.edt_elapsedTime.set_readOnly(True)
        
        # Block with time to finish
        self.lbl_timeToFinish = QLabel('Time to finish:', self)
        self.lbl_timeToFinish.setAlignment(Qt.AlignLeft)
        self.lbl_timeToFinish.setGeometry(20, 401, 150, 21)
        self.lbl_timeToFinish.setFont(QFont('Arial', 11))
        
        self.edt_timeToFinish = AdvancedQLineEdit(self)
        self.edt_timeToFinish.setAlignment(Qt.AlignRight)
        self.edt_timeToFinish.setGeometry(163, 401, 101, 21)
        self.edt_timeToFinish.setFont(QFont('Arial', 11))
        self.edt_timeToFinish.set_readOnly(True)
        
        # Block with particles generated
        self.lbl_generatedParts = QLabel('Generated particles:', self)
        self.lbl_generatedParts.setAlignment(Qt.AlignLeft)
        self.lbl_generatedParts.setGeometry(20, 424, 150, 21)
        self.lbl_generatedParts.setFont(QFont('Arial', 11))
        
        self.edt_generatedParts = AdvancedQLineEdit(self)
        self.edt_generatedParts.setAlignment(Qt.AlignRight)
        self.edt_generatedParts.setGeometry(163, 424, 101, 21)
        self.edt_generatedParts.setFont(QFont('Arial', 11))
        self.edt_generatedParts.set_readOnly(True)
        
        """Generated particle image section"""
        # Checkbox for show particle image or not
        self.chb_showParticle = QCheckBox('Show generated particle', self)
        self.chb_showParticle.setGeometry(310, 53, 200, 21)
        self.chb_showParticle.setFont(QFont('Arial', 11))
        self.chb_showParticle.clicked.connect(self.chb_showParticle_clicked)
        
        # Label with particle image
        self.lbl_particleImage = ImageLabelGenerator(self, "None")
        self.lbl_particleImage.setGeometry(310, 78, 243, 243)
        
        # Create the " particle parameters" labels, edits and units:
        dct = {
            'CEDiameter':   {'label': 'CE Diameter:',
                             'unit': '[\u03BCm]'},
            'circularity':  {'label': 'Circularity:',
                             'unit': '[-]'},
            'convexity':    {'label': 'Convexity:',
                             'unit': '[-]'},
            'elongation':   {'label': 'Elongation:',
                             'unit': '[-]'}}
        
        self.paramsLabels = {}  # Dictionary with parameters labels
        self.paramsEdits = {}  # Dictionary with parameters edits
        self.paramsUnits = {}  # Dictionary with parameters units
        
        y = 332
        for item in dct:
            self.paramsLabels[item] = QLabel(dct[item]['label'], self)
            self.paramsLabels[item].setAlignment(Qt.AlignLeft)
            self.paramsLabels[item].setGeometry(310, y, 130, 21)
            self.paramsLabels[item].setFont(QFont('Arial', 11))
            
            self.paramsEdits[item] = AdvancedQLineEdit(self)
            self.paramsEdits[item].setAlignment(Qt.AlignRight)
            self.paramsEdits[item].setGeometry(413, y, 95, 21)
            self.paramsEdits[item].setFont(QFont('Arial', 11))
            self.paramsEdits[item].set_readOnly(True)
            
            self.paramsUnits[item] = QLabel(dct[item]['unit'], self)
            self.paramsUnits[item].setAlignment(Qt.AlignLeft)
            self.paramsUnits[item].setGeometry(524, y, 50, 21)
            self.paramsUnits[item].setFont(QFont('Arial', 11))   
            y += 23

        # Buttons for start and stop generating the particles
        self.btn_generate = QPushButton('Generate', self)
        self.btn_generate.setGeometry(20, 459, 80, 27)
        self.btn_generate.clicked.connect(self.prepare_for_generation)
        
        self.btn_stopGeneration = QPushButton('Stop', self)
        self.btn_stopGeneration.setGeometry(112, 459, 70, 27)
        self.btn_stopGeneration.clicked.connect(self.stop_generation)
        
        # Timer
        self.timer = QTimer()
        self.timer.timeout.connect(self.timer_event)

        # Lamps indicator
        self.lbl_lamp = QLabel(self)
        self.lbl_lamp.setGeometry(194, 459, 25, 25)
        self.lbl_lamp.setPixmap(QPixmap('Resources/lamp_off.png'))
        
        # Progress bar
        self.progressBar = AdvancedQProgressBar(self)
        self.progressBar.setGeometry(20, 500, 534, 21)
        
        self.put_default_parameters()
        self.show()

    def put_default_parameters(self):
        """Method for putting the default values and parameters"""
        self.chb_onlySpherical.setChecked(False)
        self.onlySpherical = False
        self.nDim = 12
        self.spb_axesNum.set_value(self.nDim)
        self.picturesNum = 100
        self.spb_picturesNum.set_value(self.picturesNum)
        self.partPerPicture = 500
        self.spb_partPerPicture.set_value(self.partPerPicture)
        self.update_edt_particlesNum()
        self.PSO_nVar = self.nDim
        self.PSO_varMin = 0.0
        self.PSO_varMax = 1.0
        self.PSO_nPop = 5
        self.PSO_w = 1.0
        self.PSO_wDamp = 0.99
        self.PSO_c1 = 2.0
        self.PSO_c2 = 2.0
        self.PSO_a = 5
        self.PSO_b = 200
        self.iterLimit = 1000
        self.spb_iterLimit.set_value(self.iterLimit)
        self.precisionLimit = 0.01
        self.edt_precisionLimit.setText('{0:.4f}'.format(self.precisionLimit))
        self.useParallelSearch = False
        self.chb_useParallelSearch.setChecked(False)
        self.numThreads = 4
        self.edt_startDateTime.setText('?')
        self.elapsedTime = 0
        self.edt_elapsedTime.setText('0:00:00')
        self.edt_timeToFinish.setText('?')
        self.generatedParts = 0
        self.edt_generatedParts.setText('{0:d}'.format(self.generatedParts))
        self.percentComplete = 0
        self.progressBar.setValue(self.percentComplete)
        self.chb_showParticle.setChecked(False)
        for item in self.paramsEdits:
            self.paramsEdits[item].setText('?')
        self.btn_generate.setEnabled(False)
        self.btn_stopGeneration.setEnabled(False)

    def load_distr_data(self):
        """Method for loading the xlsx file with target parameters distributions"""
        # Determine the excel filename
        fileName = QFileDialog.getOpenFileName(self, 'Load distributions data', 'DistrData/', 'Text files (*.xlsx)')[0]
        if fileName != '':
            workbook = load_workbook(filename=fileName)
            worksheet = workbook["Data"]
            rowNum = 10  # Started row number
            # Read the xlx file and fill the arrays
            try:
                for i in range(100):
                    # Read data for CEDiameter
                    self.CEDiam_chLower[i] = worksheet.cell(row=(i + rowNum), column=3).value
                    self.CEDiam_chCentre[i] = worksheet.cell(row=(i + rowNum), column=4).value
                    self.CEDiam_chUpper[i] = worksheet.cell(row=(i + rowNum), column=5).value
                    self.init_CEDiam_distr_cum[i] = worksheet.cell(row=(i + rowNum), column=6).value
                    # Read x data for circularity, convexity and elongation
                    self.cirConEl_chLower[i] = worksheet.cell(row=(i + rowNum), column=9).value
                    self.cirConEl_chCentre[i] = worksheet.cell(row=(i + rowNum), column=10).value
                    self.cirConEl_chUpper[i] = worksheet.cell(row=(i + rowNum), column=11).value
                    # Read data for circularity
                    self.init_circ_distr_cum[i] = worksheet.cell(row=(i + rowNum), column=12).value
                    # Read data for convexity
                    self.init_convex_distr_cum[i] = worksheet.cell(row=(i + rowNum), column=18).value
                    # Read data for elongation
                    self.init_elong_distr_cum[i] = worksheet.cell(row=(i + rowNum), column=24).value
                # Finaly enable buttons for genetation
                self.btn_generate.setEnabled(True)
                self.btn_stopGeneration.setEnabled(True)
                # Enable plotting option and do not show generated plots
                self.showPlotsAct.setEnabled(True)
                self.showGeneratedPlots = False
            except:
                # Show error reading the file
                text = 'Error in reading the input xlsx file!'
                self.show_error_window(text)
            workbook.close()

    def initial_distr_treatment(self):
        """Method for the initial distributions treatment (normalization and scaling)"""
        # 1. Calculating the differential distributions
        self.init_CEDiam_distr_diff = self.calc_diff_from_cum(self.init_CEDiam_distr_cum)
        self.init_circ_distr_diff = self.calc_diff_from_cum(self.init_circ_distr_cum)
        self.init_convex_distr_diff = self.calc_diff_from_cum(self.init_convex_distr_cum)
        self.init_elong_distr_diff = self.calc_diff_from_cum(self.init_elong_distr_cum)

        # 2. Normalization the distributions
        self.norm_CEDiam_distr_diff = self.normalize_diff(self.init_CEDiam_distr_diff)
        self.norm_circ_distr_diff = self.normalize_diff(self.init_circ_distr_diff)
        self.norm_convex_distr_diff = self.normalize_diff(self.init_convex_distr_diff)
        self.norm_elong_distr_diff = self.normalize_diff(self.init_elong_distr_diff)

        # 3. Slicing the distribution (determination boundaries of non 0 values)
        (self.CEDiam_leftBndChannel, self.CEDiam_rightBndChannel) = self.calc_boundaries(self.norm_CEDiam_distr_diff)
        (self.circ_leftBndChannel, self.circ_rightBndChannel) = self.calc_boundaries(self.norm_circ_distr_diff)
        (self.convex_leftBndChannel, self.convex_rightBndChannel) = self.calc_boundaries(self.norm_convex_distr_diff)
        (self.elong_leftBndChannel, self.elong_rightBndChannel) = self.calc_boundaries(self.norm_elong_distr_diff)

    def calc_diff_from_cum(self, cum):
        """Method for producing the differential distribution from the cumulative one"""
        diff = np.zeros(100)
        diff[0] = cum[0]
        for i in range(1, 100):
            diff[i] = cum[i] - cum[i - 1]
            if diff[i] < 0:
                diff[i] = 0  # For safety
        return diff
        
    def normalize_diff(self, diff):
        """Method for normalizing the differential distribution"""
        # Normalization procedure
        normDiff = np.zeros(100)
        maxNum = np.max(diff)
        for i in range(100):
            normDiff[i] = diff[i] / maxNum
        return normDiff

    def calc_boundaries(self, normDiff):
        """Method for determining the differential distribution boundaries"""
        leftBndChannel = 0
        for i in range(100):  # Determining the left boundary
            if normDiff[i] > 0:
                leftBndChannel = i
                break
        rightBndChannel = 99
        for i in range(100):  # Determining the right boundary
            if normDiff[99 - i] > 0:
                rightBndChannel = 99 - i
                break
        return leftBndChannel, rightBndChannel

    def show_distr_plots(self):
        """Method for show loaded distributions plots"""
        # Creating plot in memory
        fig1 = plt.figure(1)
        fig1.canvas.set_window_title('Particle properties distributions')
        fig1.clear()

        # Plotting CEDiameter distribution
        axCEDiam = fig1.add_subplot(221)
        axCEDiam.set_xscale('log')
        axCEDiam.set_xlabel(r'CE Diameter, [$\mu m$]')
        axCEDiam.set_ylabel(r'Cumulative, [%]')
        line = axCEDiam.plot(self.CEDiam_chUpper, self.init_CEDiam_distr_cum, label='Initial')
        plt.setp(line, color='b')
        if self.showGeneratedPlots:
            line = axCEDiam.plot(self.CEDiam_chUpper, self.gen_CEDiam_distr_cum, label='Generated')
            plt.setp(line, color='r')
            axCEDiam.legend(bbox_to_anchor=(0.0, 1.05), loc='lower left', borderaxespad=0.)

        # Plotting Circularity distribution
        axCirc = fig1.add_subplot(222)
        axCirc.set_xlabel(r'Circularity, [-]')
        axCirc.set_ylabel(r'Cumulative, [%]')
        line = axCirc.plot(self.cirConEl_chUpper, self.init_circ_distr_cum, label='Initial')
        plt.setp(line, color='b')
        if self.showGeneratedPlots and not self.onlySpherical:
            line = axCirc.plot(self.cirConEl_chUpper, self.gen_circ_distr_cum, label='Generated')
            plt.setp(line, color='r')

        # Plotting Convexity distribution
        axConvex = fig1.add_subplot(223)
        axConvex.set_xlabel(r'Convexity, [-]')
        axConvex.set_ylabel(r'Cumulative, [%]')
        line = axConvex.plot(self.cirConEl_chUpper, self.init_convex_distr_cum, label='Initial')
        plt.setp(line, color='b')
        if self.showGeneratedPlots and not self.onlySpherical:
            line = axConvex.plot(self.cirConEl_chUpper, self.gen_convex_distr_cum, label='Generated')
            plt.setp(line, color='r')

        # Plotting Elongation distribution
        axElong = fig1.add_subplot(224)
        axElong.set_xlabel(r'Elongation, [-]')
        axElong.set_ylabel(r'Cumulative, [%]')
        line = axElong.plot(self.cirConEl_chUpper, self.init_elong_distr_cum, label='Initial')
        plt.setp(line, color='b')
        if self.showGeneratedPlots and not self.onlySpherical:
            line = axElong.plot(self.cirConEl_chUpper, self.gen_elong_distr_cum, label='Generated')
            plt.setp(line, color='r')

        plt.subplots_adjust(top=0.92, bottom=0.10, left=0.10, right=0.95, hspace=0.40, wspace=0.30)
        fig1.canvas.draw()  # Update the canvas
        plt.show()  # Show the plots

    def chb_onlySpherical_clicked(self):
        """Method to define weather generate only spherical particles or not"""
        if self.chb_onlySpherical.isChecked():
            self.onlySpherical = True
            if self.chb_showParticle.isChecked():
                self.lbl_particleImage.drawFlag = "Circ"
            else:
                self.lbl_particleImage.drawFlag = "None"
            self.useParallelSearch = False
            self.chb_useParallelSearch.setChecked(False)
        else:
            self.onlySpherical = False
            if self.chb_showParticle.isChecked():
                self.lbl_particleImage.drawFlag = "Part"
            else:
                self.lbl_particleImage.drawFlag = "None"
        
        # Hide or show specific spin boxes, edits and buttons
        flag = not self.onlySpherical
        self.chb_useParallelSearch.setEnabled(flag)
        self.spb_axesNum.setEnabled(flag)
        self.btn_resetAxesNum.setEnabled(flag)
        self.btn_PSOAlgSettings.setEnabled(flag)
        self.spb_iterLimit.setEnabled(flag)
        self.btn_resetIterLimit.setEnabled(flag)
        self.edt_precisionLimit.setEnabled(flag)
        self.btn_resetPrecisionLimit.setEnabled(flag)
        self.btn_ParallelSearchSettings.setEnabled(flag)
        self.update()

    def chb_showParticle_clicked(self):
        """Method to define weather show the generated particle or not"""
        if self.chb_showParticle.isChecked():
            if self.chb_onlySpherical.isChecked():
                self.lbl_particleImage.drawFlag = "Circ"
            elif self.chb_useParallelSearch.isChecked():
                self.lbl_particleImage.drawFlag = "PSImg"
            else:
                self.lbl_particleImage.drawFlag = "Part"
        else:
            self.lbl_particleImage.drawFlag = "None"
        self.update()

    def chb_useParallelSearch_clicked(self):
        """Method to define weather to use the parallel search or not"""
        if self.chb_useParallelSearch.isChecked():
            self.useParallelSearch = True
            if self.chb_showParticle.isChecked():
                self.lbl_particleImage.drawFlag = "PSImg"
            else:
                self.lbl_particleImage.drawFlag = "None"
            self.chb_onlySpherical.setEnabled(False)
        else:
            self.useParallelSearch = False
            self.chb_onlySpherical.setEnabled(True)
            if self.chb_showParticle.isChecked():
                self.lbl_particleImage.drawFlag = "Part"
            else:
                self.lbl_particleImage.drawFlag = "None"
        self.update()

    def val_changed_spb_axesNum(self):
        """Method for change spb_axesNum value"""
        self.nDim = self.spb_axesNum.value()
        self.PSO_nVar = self.nDim
        self.update()
    
    def update_edt_particlesNum(self):
        """Method for update the edit with total particles number"""
        self.picturesNum = self.spb_picturesNum.value()
        self.partPerPicture = self.spb_partPerPicture.value()
        self.particlesNum = self.picturesNum * self.partPerPicture
        self.edt_particlesNum.setText('{0:d}'.format(self.particlesNum))    
    
    def val_changed_spb_iterLimit(self):
        """Method for change spb_iterLimit value"""
        self.iterLimit = self.spb_iterLimit.value()
    
    def update_precisionLimit(self):
        oldValue = self.precisionLimit  # Save old value
        editStr = self.edt_precisionLimit.text()
        (newValue, errorCode, errorText) = edit_str_to_value('Precision limit', editStr)
        if errorCode == 0:  # No errors
            self.precisionLimit = newValue
            self.edt_precisionLimit.setText('{0:.4f}'.format(newValue))
        else:  # Error
            self.edt_precisionLimit.setText('{0:.4f}'.format(oldValue))
            self.show_error_window(errorText)

    def reset_axesNum(self):
        """Method for reset the spb_axesNum value to default"""
        self.nDim = 12
        self.spb_axesNum.setValue(self.nDim)

    def reset_picturesNum(self):
        """Method for reset the spb_picturesNum value to default"""
        self.picturesNum = 100
        self.spb_picturesNum.set_value(self.picturesNum)
        self.update_edt_particlesNum()

    def reset_partPerPicture(self):
        """Method for reset the spb_partPerPicture value to default"""
        self.partPerPicture = 500
        self.spb_partPerPicture.set_value(self.partPerPicture)
        self.update_edt_particlesNum()

    def reset_iterLimit(self):
        self.iterLimit = 1000
        self.spb_iterLimit.set_value(self.iterLimit)
        
    def reset_precisionLimit(self):
        self.precisionLimit = 0.01
        self.edt_precisionLimit.setText('{0:.4f}'.format(self.precisionLimit))

    def open_PSOAlg_settings(self):
        """Method to open window with PSO algorithm settings"""
        self.settingsWindow = PSOSettingsWindow(self)
        
    def open_ParallelSearch_settings(self):
        """Method to open window with parallel search settings"""
        self.pSearchSettingsWindow = PSearchSettingsWindow(self)

    def set_PSO_settings_data(self, settingsData):
        """Method for save setings data comming from settings window"""
        # Update the PSO algorithm parameters
        self.PSO_nPop = settingsData['PSO_nPop']
        self.PSO_w = settingsData['PSO_w']
        self.PSO_wDamp = settingsData['PSO_wDamp']
        self.PSO_c1 = settingsData['PSO_c1']
        self.PSO_c2 = settingsData['PSO_c2']
        self.PSO_a = settingsData['PSO_a']
        self.PSO_b = settingsData['PSO_b']
        
    def set_PSearch_settings_data(self, settingsData):
        """Method for save the PSearch settings data"""
        self.numThreads = settingsData['numThreads']

    def prepare_for_generation(self):
        """Method for initial preparation for the generation process"""
        # Choose the output text file with generated particles data
        self.fileName = QFileDialog.getSaveFileName(self, 'Save generated particles system',
                                               'GenPartSystems/untitled.txt', 'Text files (*.txt)')[0]
        if self.fileName != '':
            with open(self.fileName, 'w') as outfile:
                # Write some basic information about the particles system
                outfile.write('{0}\n'.format(int(self.onlySpherical)))
                outfile.write('{0}\n'.format(self.nDim))
                outfile.write('{0}\n'.format(self.picturesNum))
                outfile.write('{0}\n'.format(self.partPerPicture))
                self.change_lamp_state()  
        else:
            text = 'Output file has not been chosen.\nChoose the file!'
            self.show_error_window(text)
            return None  # Exit the generation method
        
        # Set of preparations:
        self.percentComplete = 0  # Set the progress bar to 0%
        self.progressBar.setValue(self.percentComplete)
        self.enable_elements(False)  # Disable some window elements      
        self.timeToFinish = 0
        self.sumAreaUm2 = 0  # Zero the sum of particles areas
        self.elapsedTime = 0  # Reset the elapsed time
        self.timer.start(1000)
        self.stopGeneration = False  # Do the generation
        
        # Save the generation started date-time
        dateTime = strftime("%Y-%m-%d %H:%M:%S", localtime())
        self.edt_startDateTime.setFont(QFont('Arial', 7))
        self.edt_startDateTime.setText(dateTime)        
        
        # Use default or parallel generation procedure
        if self.useParallelSearch:
            self.make_parallel_generation_do_before()
        else:
            self.make_generation_do_before()
    
    """========== START set of methods for default generation process =========="""
    def make_generation_do_before(self):
        """Preparation method for start generation of the particles"""
        # Make initial distribution treatment
        self.initial_distr_treatment() # Make the initial distributions treatment
        
        # Empty the count arrays for further calc diff and cum distributions
        self.count_CEDiam_distr_diff = np.zeros(100)
        self.count_circ_distr_diff = np.zeros(100)
        self.count_convex_distr_diff = np.zeros(100)
        self.count_elong_distr_diff = np.zeros(100)
        self.count_solid_distr_diff = np.zeros(100) 
        
        # Make worker with multi threading
        worker = Worker(self.make_generation_main_process)
        worker.signals.progress.connect(self.update_params_default_generation)
        worker.signals.finished.connect(self.make_generation_do_after)
        self.threadpool.start(worker)
        
    def make_generation_main_process(self, progress_callback):
        """Main method for the default particles generation"""
        progressData = {'dims': None,  # Recently found particle dims
                        'CEDiameter': None,  # Recently found particle CEDiameter
                        'circularity': None,  # Recently found particle circularity
                        'convexity': None,  # Recently found particle convexity
                        'elongation': None,  # Recently found particle elongation
                        'percentComplete': None,  # Percentage to complete the generation
                        'timeToFinish': None,  # Time to finish as formated string
                        'genParticles': None}  # Iteration number

        # Open output file to save the generated particles
        outfile = open(self.fileName, 'a')

        # Main particle generation loop:
        for i in range(self.particlesNum):
            # Check to stop the generation
            if self.stopGeneration:
                break
            
            # Calculate the time to finish
            if i == 0:
                self.genStartTime = time()  # Start time
                timeString = '00:00:00'
            elif i % 20 == 0:
                timeDelta = time() - self.genStartTime
                self.genStartTime = time()
                value = ((self.particlesNum - i - 1) // 20) * timeDelta
                self.timeToFinish = int(round((8 * self.timeToFinish + 2 * value) / 10))  # Damping for smoothing
                timeString = self.make_label_for_time(self.timeToFinish)
            progressData['timeToFinish'] = timeString
            
            # Generation the desired parameters from the distribution:
            if self.onlySpherical:
                self.gen_CEDiameter = self.get_value_from_distribution('CEDiam')
            else:
                self.target_CEDiameter = self.get_value_from_distribution('CEDiam')
                self.target_circularity = self.get_value_from_distribution('circ')
                self.target_convexity = self.get_value_from_distribution('convex')
                self.target_elongation = self.get_value_from_distribution('elong')                
            
            if self.onlySpherical:
                self.sumAreaUm2 += (m.pi * m.pow(self.gen_CEDiameter, 2)) / 4
            else:
                # Search for the shape of particle with desired parameters:
                # Execute the function for the search   
                results = self.psoAlg_dll.run_search(
                    init_circularity = self.target_circularity,
                    init_convexity = self.target_convexity,
                    init_elongation = self.target_elongation,
                    nVar = self.PSO_nVar,
                    varMin = self.PSO_varMin,
                    varMax = self.PSO_varMax,
                    useIterLimit = True,
                    iterLimit = self.iterLimit,
                    usePrecisionLimit = True,
                    precisionLimit = self.precisionLimit,
                    showErrorPlot = False,
                    nPop = self.PSO_nPop,
                    w = self.PSO_w,
                    wDamp = self.PSO_wDamp,
                    c1 = self.PSO_c1,
                    c2 = self.PSO_c2,
                    a = self.PSO_a,
                    b = self.PSO_b)
                          
                # Determine the found particle parameters:
                self.gen_dims = results['globalBestPosition']
                result_params = self.particle.get_particle_parameters(1.0, self.gen_dims, self.nDim)
                areaPixels = result_params['areaPixels']
                imgScale = self.target_CEDiameter * m.sqrt(m.pi / (areaPixels * 4))
                result_params = self.particle.get_particle_parameters(imgScale, self.gen_dims, self.nDim)
            
                self.gen_CEDiameter = result_params['CEDiameter']
                self.gen_circularity = result_params['circularity']
                self.gen_convexity = result_params['convexity']
                self.gen_elongation = result_params['elongation']
                self.gen_solidity = result_params['solidity']
                self.sumAreaUm2 += result_params['areaUm2']
            
            # Update the percent complete for the progress bar:
            self.percentComplete = i * 100 / (self.particlesNum - 1)
            progressData['percentComplete'] = self.percentComplete

            # Update the count arrays:
            self.update_count_array(self.gen_CEDiameter, self.CEDiam_chLower, 
                self.CEDiam_chUpper, self.count_CEDiam_distr_diff)
            if not self.onlySpherical:
                self.update_count_array(self.gen_circularity, self.cirConEl_chLower,
                    self.cirConEl_chUpper, self.count_circ_distr_diff)
                self.update_count_array(self.gen_convexity, self.cirConEl_chLower,
                    self.cirConEl_chUpper, self.count_convex_distr_diff)
                self.update_count_array(self.gen_elongation, self.cirConEl_chLower,
                    self.cirConEl_chUpper, self.count_elong_distr_diff)
                self.update_count_array(self.gen_solidity, self.cirConEl_chLower,
                    self.cirConEl_chUpper, self.count_solid_distr_diff)
            
            # Update the progressData dictionary:
            progressData['genParticles'] = i + 1
            progressData['CEDiameter'] = self.gen_CEDiameter
            if not self.onlySpherical:
                progressData['dims'] = self.gen_dims 
                progressData['circularity'] = self.gen_circularity
                progressData['convexity'] = self.gen_convexity
                progressData['elongation'] = self.gen_elongation
                    
            # Save the current particle data to the output file:
            if self.onlySpherical:
                outfile.write('{0:d},{1:.5f}'.format(i, self.gen_CEDiameter))
            else:
                outfile.write('{0:d},{1:.5f}'.format(i, imgScale))
                for k in range(self.nDim):
                    outfile.write(',{0:.5f}'.format(self.gen_dims[k]))
            outfile.write('\n')  # Go to the new line

            # Add the calculates sumAreaUm2 to the end of the file
            if i == (self.particlesNum - 1):
                outfile.write('{0:.3f}'.format(self.sumAreaUm2))

            # Send the callback
            progress_callback.emit(progressData)

        # Close the output file
        outfile.close()
        
    def update_params_default_generation(self, progressData):
        """Method to update particle shape and parameters during the generation"""
        # Update the image of last found particle
        if not self.onlySpherical:
            self.lbl_particleImage.set_dimsValues(progressData['dims'])
        
        # Update the particle parameters
        self.paramsEdits['CEDiameter'].setText('{0:.2f}'.format(progressData['CEDiameter']))
        if self.onlySpherical:
            self.paramsEdits['circularity'].setText('?')
            self.paramsEdits['convexity'].setText('?')
            self.paramsEdits['elongation'].setText('?')
        else:   
            self.paramsEdits['circularity'].setText('{0:.3f}'.format(progressData['circularity']))
            self.paramsEdits['convexity'].setText('{0:.3f}'.format(progressData['convexity']))
            self.paramsEdits['elongation'].setText('{0:.3f}'.format(progressData['elongation']))       
            
        # Determine the average time to complete the generation:
        self.edt_timeToFinish.setText(progressData['timeToFinish'])

        # Update the number of generated particles
        self.edt_generatedParts.setText('{0:d}'.format(progressData['genParticles']))

        # Update the progress bar
        self.progressBar.setValue(progressData['percentComplete'])
        self.update()

    def make_generation_do_after(self):
        """Method for doing some things after the default generation procedure"""     
        # Calculate the generated differential and cumulative distributions
        self.gen_CEDiam_distr_diff, self.gen_CEDiam_distr_cum = \
            self.make_distr_from_count_array('CEDiam')
        if not self.onlySpherical:
            self.gen_circ_distr_diff, self.gen_circ_distr_cum = \
                self.make_distr_from_count_array('circ')
            self.gen_convex_distr_diff, self.gen_convex_distr_cum = \
                self.make_distr_from_count_array('convex')
            self.gen_elong_distr_diff, self.gen_elong_distr_cum = \
                self.make_distr_from_count_array('elong')
            self.gen_solid_distr_diff, self.gen_solid_distr_cum = \
                self.make_distr_from_count_array('solid')   
                
        self.change_lamp_state()  # Change the state of the green lamp
        self.edt_timeToFinish.setText('00:00:00')
        self.timer.stop()  # Stop the timer
        self.enable_elements(True)  # Enable window elements
        self.showGeneratedPlots = True  # The generated plots can be shown
        self.make_output_xlsx_file()
        
        if self.stopGeneration:
            text = 'Generation has been stopped!'
        else:
            text = 'Generation of particles system is finished!'
        self.show_information_window(text) 
        
    """========== END set of methods for default generation process ==========""" 
    
    
    """========== START set of methods for parallel generation process ==========""" 
    def make_parallel_generation_do_before(self):
        """Preparation method for start parallel generation of the particles"""
        
        # Clarify the real number of parallel threads
        if self.particlesNum < self.numThreads:
            self.numThreads = self.particlesNum
            
        # Make worker with multi threading
        worker = Worker(self.make_parallel_generation_main_process)
        worker.signals.progress.connect(self.update_params_parallel_generation)
        worker.signals.finished.connect(self.make_parallel_generation_do_after)
        self.threadpool.start(worker)

    def make_parallel_generation_main_process(self, progress_callback):
        """Main method for the parallel particles generation"""
        
        progressData = {'percentComplete': None}  # Percentage to complete the generation
                        
        # Empty everything in "Modules/generator_c/data/" folder
        # Empty everything in "Modules/generator_c/run/" folder
        self.deleteError = False
        folder = os.getcwd() + r"\Modules\Generator_c\data"
        for the_file in os.listdir(folder):
            file_path = os.path.join(folder, the_file)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except:
                self.deleteError = True
        
        folder = os.getcwd() + r"\Modules\Generator_c\run"
        for the_file in os.listdir(folder):
            file_path = os.path.join(folder, the_file)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except:
                self.deleteError = True
        
        # Make txt files with initial parameters distributions (1 for every thread)
        if not self.deleteError:
            folder = os.getcwd() + r"\Modules\Generator_c\data"
            for i in range(1, self.numThreads + 1):
                fileName = folder + "\init_params_distr_{0}.txt".format(i)
                textFile = open(fileName, 'w')
                
                # Write data to the opened txt file
                for j in range(0, 100):
                    # Channel number
                    textFile.write("{0:d} ".format(j + 1))
                    
                    # CEDiameter initial distribution data
                    textFile.write("{0:.3f} ".format(self.CEDiam_chLower[j]))
                    textFile.write("{0:.3f} ".format(self.CEDiam_chCentre[j]))
                    textFile.write("{0:.3f} ".format(self.CEDiam_chUpper[j]))
                    textFile.write("{0:.3f} ".format(self.init_CEDiam_distr_cum[j]))
                    
                    # cirConEl channel numbers (common for circ, convex and elong)
                    textFile.write("{0:.3f} ".format(self.cirConEl_chLower[j]))
                    textFile.write("{0:.3f} ".format(self.cirConEl_chCentre[j]))
                    textFile.write("{0:.3f} ".format(self.cirConEl_chUpper[j]))
                    
                    # Circularity, convexity and elongation init distr data
                    textFile.write("{0:.3f} ".format(self.init_circ_distr_cum[j]))
                    textFile.write("{0:.3f} ".format(self.init_convex_distr_cum[j]))
                    textFile.write("{0:.3f} ".format(self.init_elong_distr_cum[j]))
                    textFile.write("\n")
                textFile.close()

        # Calculate the amount of particles to be generated by each thread
        partPerThread = [];
        temp_1 = self.particlesNum // self.numThreads
        for i in range(self.numThreads):
            if i == (self.numThreads - 1):
                temp_2 = (self.particlesNum - self.numThreads * temp_1) + temp_1
                partPerThread.append(temp_2)
            else:
                partPerThread.append(temp_1)
        
        # Make .bat files to run the generators
        if not self.deleteError:
            folder = os.getcwd() + r"\Modules\Generator_c\run"
            for i in range(1, self.numThreads + 1):
                fileName = folder + "\Run_{0}.bat".format(i)
                textFile = open(fileName, 'w')
                
                # Write data to the opened txt file
                textFile.write(r"@echo off")
                textFile.write("\n")
                textFile.write("pushd \"{0}\"\n".format(folder))
                textFile.write(r"cd ..\build")
                textFile.write("\n")
                
                useIterLimit = 1
                usePrecisionLimit = 1
                showErrorPlot = 1
                progStr = "generator_c.exe" + \
                          " {0:d}".format((i)) + \
                          " {0:d}".format(partPerThread[i - 1]) + \
                          " {0:d}".format(self.PSO_nVar) + \
                          " {0:f}".format(self.PSO_varMin) + \
                          " {0:f}".format(self.PSO_varMax) + \
                          " {0:d}".format(useIterLimit) + \
                          " {0:d}".format(self.iterLimit) + \
                          " {0:d}".format(usePrecisionLimit) + \
                          " {0:f}".format(self.precisionLimit) + \
                          " {0:d}".format(showErrorPlot) + \
                          " {0:d}".format(self.PSO_nPop) + \
                          " {0:f}".format(self.PSO_w) + \
                          " {0:f}".format(self.PSO_wDamp) + \
                          " {0:f}".format(self.PSO_c1) + \
                          " {0:f}".format(self.PSO_c2) + \
                          " {0:d}".format(self.PSO_a) + \
                          " {0:d}".format(self.PSO_b)
                textFile.write(progStr)
            textFile.close()
             
        # Send the callback
        progressData['percentComplete'] = 30
        progress_callback.emit(progressData)
        
        # Run every .bat file one by one
        folder = os.getcwd() + r"\Modules\Generator_c\run"
        for i in range(1, self.numThreads + 1):
            fileName = "\"" + folder + "\Run_{0}.bat".format(i) + "\""
            Popen(fileName, creationflags=CREATE_NEW_CONSOLE)
            sleep(3)
        
        # Loop for waiting for the end of particles generation
        searchNext = True
        folder = os.getcwd() + r"\Modules\Generator_c\data"
        while searchNext:
            finished = 0
            for i in range(1, self.numThreads + 1):
                fileName = folder + "\generated_info_{0}.txt".format(i)
                if os.path.isfile(fileName):
                    finished += 1
            if finished == self.numThreads:
                searchNext = False
            sleep(3)
            
        # Send the callback
        progressData['percentComplete'] = 60
        progress_callback.emit(progressData)
        
        # Compilation of the generated cumulative distributions for the plots
        self.gen_CEDiam_distr_cum = np.zeros(100)
        self.gen_circ_distr_cum = np.zeros(100)
        self.gen_convex_distr_cum = np.zeros(100) 
        self.gen_elong_distr_cum = np.zeros(100)
        self.gen_solid_distr_cum = np.zeros(100)
        partPerThread = [];  # Final amount of generated particles per thread
        self.sumAreaUm2 = 0  # Final summ of particles areas
        
        folder = os.getcwd() + r"\Modules\Generator_c\data"
        for i in range(self.numThreads):
            fileName = folder + "\generated_info_{0}.txt".format(i + 1)
            
            # Open the file for read the data
            tempFile = open(fileName, 'r')
            
            # Read and store the number of particles per thread and summ of areas
            lineStr = tempFile.readline()
            partPerThread.append(int(lineStr))
            lineStr = tempFile.readline()
            self.sumAreaUm2 += float(lineStr)
            
            # Read the cumulative distributions data
            for j in range(100):
                lineStr = tempFile.readline()
                listStr = lineStr.split(',')
                listNum = [float(x) for x in listStr]
                
                # Update the cumulative distribution arrays
                self.gen_CEDiam_distr_cum[j] += listNum[0]
                self.gen_circ_distr_cum[j] += listNum[1]
                self.gen_convex_distr_cum[j] += listNum[2]
                self.gen_elong_distr_cum[j] += listNum[3]
                self.gen_solid_distr_cum[j] += listNum[4]
            
            tempFile.close()
            
        # Finilize the calculation of the cumulative distributions
        for i in range(100):
            temp = self.gen_CEDiam_distr_cum[i] / self.numThreads
            self.gen_CEDiam_distr_cum[i] = temp
            temp = self.gen_circ_distr_cum[i] / self.numThreads
            self.gen_circ_distr_cum[i] = temp
            temp = self.gen_convex_distr_cum[i] / self.numThreads
            self.gen_convex_distr_cum[i] = temp
            temp = self.gen_elong_distr_cum[i] / self.numThreads
            self.gen_elong_distr_cum[i] = temp          
            temp = self.gen_solid_distr_cum[i] / self.numThreads
            self.gen_solid_distr_cum[i] = temp

        # Compilation of the generated particles data into one file
        folder = os.getcwd() + r"\Modules\Generator_c\data"
        particleNum = 0
        outfile = open(self.fileName, 'a')  # Main file with generated data
        for i in range(self.numThreads):
            tempFileName = folder + "\generated_data_{0}.txt".format(i + 1)
            tempFile = open(tempFileName, 'r')
            
            # Read the every row data from temp file
            for j in range(partPerThread[i]):
                lineStr = tempFile.readline()
                listStr = lineStr.split(',')
                
                # Replace the first element with overal particleNum
                listStr[0] = "{0:d}".format(particleNum)
                particleNum += 1
                lineStr = ",".join(listStr) 
                
                # Write the new line to the output file
                outfile.write(lineStr)

        # Finaly write the value of particles sum srea
        outfile.write('{0:.3f}'.format(self.sumAreaUm2))
        outfile.close()
           
    def update_params_parallel_generation(self, progressData):
        # Update the progress bar
        self.progressBar.setValue(progressData['percentComplete'])
        self.update()
    
    def make_parallel_generation_do_after(self):
        """Method for doing some things after the parallel generation procedure"""
        # Read the generated differential and cumulative distributions
        # ...
        
        self.change_lamp_state()  # Change the state of the green lamp
        self.edt_timeToFinish.setText('00:00:00')
        self.timer.stop()  # Stop the timer
        self.enable_elements(True)  # Enable window elements
        
        if self.deleteError:
            text = 'Error while deleting the files in "data" folder!'
        else:
            self.showGeneratedPlots = True  # The generated plots can be shown
            self.make_output_xlsx_file()
            if self.stopGeneration:
                text = 'Generation has been stopped!'
            else:
                text = 'Generation of particles system is finished!'
            
        self.progressBar.setValue(100)
        self.show_information_window(text)
        
    """========== END set of methods for parallel generation process =========="""
  
    def stop_generation(self):
        """Method for interupting the generation"""
        if self.useParallelSearch:
            folder = os.getcwd() + r"\Modules\Generator_c\data"
            fileName = folder + r"\stop.txt"
            tempFile = open(fileName, 'w')
            tempFile.close()
        self.stopGeneration = True

    def get_value_from_distribution(self, distrName):
        """Method to generate a value from the specific distribution"""
        # Determine the distribution and its parameters
        if distrName == 'CEDiam':
            normDiff = self.norm_CEDiam_distr_diff
            chLower = self.CEDiam_chLower
            chUpper = self.CEDiam_chUpper     
            leftBndChannel = self.CEDiam_leftBndChannel
            rightBndChannel = self.CEDiam_rightBndChannel
            leftBndValue = m.log(chLower[leftBndChannel], 10)
            rightBndValue = m.log(chUpper[rightBndChannel], 10)
        elif distrName == 'circ':
            normDiff = self.norm_circ_distr_diff
            chLower = self.cirConEl_chLower
            chUpper = self.cirConEl_chUpper
            leftBndChannel = self.circ_leftBndChannel
            rightBndChannel = self.circ_rightBndChannel
            leftBndValue = chLower[leftBndChannel]
            rightBndValue = chUpper[rightBndChannel]
        elif distrName == 'convex':
            normDiff = self.norm_convex_distr_diff
            chLower = self.cirConEl_chLower
            chUpper = self.cirConEl_chUpper
            leftBndChannel = self.convex_leftBndChannel
            rightBndChannel = self.convex_rightBndChannel
            leftBndValue = chLower[leftBndChannel]
            rightBndValue = chUpper[rightBndChannel]
        elif distrName == 'elong':
            normDiff = self.norm_elong_distr_diff
            chLower = self.cirConEl_chLower
            chUpper = self.cirConEl_chUpper
            leftBndChannel = self.elong_leftBndChannel
            rightBndChannel = self.elong_rightBndChannel
            leftBndValue = chLower[leftBndChannel]
            rightBndValue = chUpper[rightBndChannel]
        
        # Make search for the value from the distribution
        doSearch = True
        while doSearch:
            xTry = rnd.uniform(leftBndValue, rightBndValue)
            if distrName == 'CEDiam':
                xTry = m.pow(10, xTry)     
            prob = 1.0
            # Determine the channel number
            for i in range(0, 100):
                if (xTry >= chLower[i]) and (xTry <= chUpper[i]):
                    prob = normDiff[i]
                    break
            # Accept the particle or not according to probability
            rndNumber = rnd.uniform(0.0, 1.0)  # 0.0 <= rndNumber < 1.0
            if rndNumber < prob:  # Particle accepted
                return xTry
    
    @staticmethod
    def update_count_array(value, chLower, chUpper, countDistr):
        """Method for update the count array in every iteration"""
        for i in range(100):
            if (value >= chLower[i]) and (value <= chUpper[i]):
                countDistr[i] += 1
                break

    def make_distr_from_count_array(self, param):
        """Method for making diff and cum disributions from the count array"""
        if param == 'CEDiam':
            countDistr = self.count_CEDiam_distr_diff
        elif param == 'circ':
            countDistr = self.count_circ_distr_diff
        elif param == 'convex':
            countDistr = self.count_convex_distr_diff
        elif param == 'elong':
            countDistr = self.count_elong_distr_diff
        elif param == 'solid':
            countDistr = self.count_solid_distr_diff
        
        partSum = np.sum(countDistr)
        # make differential distribution
        outputDistr_diff = np.zeros(100)
        for i in range(0, 100):
            outputDistr_diff[i] = countDistr[i]*100 / partSum
        # Make cumulative distribution
        outputDistr_cum = np.zeros(100)
        outputDistr_cum[0] = outputDistr_diff[0]
        for i in range(1, 100):
            outputDistr_cum[i] = outputDistr_cum[i - 1] + outputDistr_diff[i]
        return outputDistr_diff, outputDistr_cum
        
    def change_lamp_state(self):
        """Method for change the lamp state"""
        if self.lamp:
            self.lbl_lamp.setPixmap(QPixmap('Resources/lamp_off.png'))
            self.lamp = False
        else:
            self.lbl_lamp.setPixmap(QPixmap('Resources/lamp_on.png'))
            self.lamp = True
        
    def enable_elements(self, flag):
        """Method for enable or disable some elements"""
        self.loadDistrAct.setEnabled(flag)
        self.chb_onlySpherical.setEnabled(flag)
        self.spb_axesNum.setEnabled(flag)
        self.btn_resetAxesNum.setEnabled(flag)
        self.spb_picturesNum.setEnabled(flag)
        self.btn_resetPicturesNum.setEnabled(flag)
        self.spb_partPerPicture.setEnabled(flag)
        self.btn_resetPartPerPicture.setEnabled(flag)
        self.btn_PSOAlgSettings.setEnabled(flag)
        self.spb_iterLimit.setEnabled(flag)
        self.btn_resetIterLimit.setEnabled(flag)
        self.edt_precisionLimit.setEnabled(flag)
        self.btn_resetPrecisionLimit.setEnabled(flag)
        self.btn_generate.setEnabled(flag)

    def make_output_xlsx_file(self):
        """Method for creation and saving the output xlsx file"""
        workbook = Workbook()
        worksheet = workbook.create_sheet("Data", 0) # insert at first position
        
        # Adjust width of some columns
        array = ('A', 'G', 'M', 'S', 'Y')
        for item in array:
            worksheet.column_dimensions[item].width = 2.85
        
        # Title
        worksheet['B2'] = 'Generated parameters distributions'
        
        # Merge necessary cells
        array = ('B4:F4', 'H4:L4', 'N4:R4', 'T4:X4', 'Z4:AD4')
        for item in array:
            worksheet.merge_cells(item)

        # Table names:
        array = (('B4', 'CEDiameters distribution'), ('H4', 'Circularity distribution'),
                 ('N4', 'Convexity distribution'), ('T4', 'Elongation distribution'),
                 ('Z4', 'Solidity distribution'))
        for cell, title in array:
            worksheet[cell] = title
        
        # Table titles:
        array_1 = ('Ch N', 'Ch low', 'Ch cent', 'Ch upp', 'Cum N')
        array_2 = ('[-]', '[-]', '[-]', '[-]', '[%]')
        array_3 = ('[-]', '[um]', '[um]', '[um]', '[%]')
        colNum = 2
        for i in range(5):
            for j in range(5):   
                worksheet.cell(row=5, column=colNum).value = array_1[j]
                if i == 0:
                    worksheet.cell(row=6, column=colNum).value = array_3[j]
                else:
                    worksheet.cell(row=6, column=colNum).value = array_2[j]
                colNum += 1
            colNum += 1
            

        # Fill the tables
        rowNum = 7
        for i in range(100):
            # CEDiameter
            worksheet.cell(row=rowNum, column=2).value = i + 1
            worksheet.cell(row=rowNum, column=3).value = self.CEDiam_chLower[i]
            worksheet.cell(row=rowNum, column=4).value = self.CEDiam_chCentre[i]
            worksheet.cell(row=rowNum, column=5).value = self.CEDiam_chUpper[i]
            worksheet.cell(row=rowNum, column=6).value = self.gen_CEDiam_distr_cum[i]
            if not self.onlySpherical:
                # Circularity
                worksheet.cell(row=rowNum, column=8).value = i + 1
                worksheet.cell(row=rowNum, column=9).value = self.cirConEl_chLower[i]
                worksheet.cell(row=rowNum, column=10).value = self.cirConEl_chCentre[i]
                worksheet.cell(row=rowNum, column=11).value = self.cirConEl_chUpper[i]
                worksheet.cell(row=rowNum, column=12).value = self.gen_circ_distr_cum[i]
                # Convexity
                worksheet.cell(row=rowNum, column=14).value = i + 1
                worksheet.cell(row=rowNum, column=15).value = self.cirConEl_chLower[i]
                worksheet.cell(row=rowNum, column=16).value = self.cirConEl_chCentre[i]
                worksheet.cell(row=rowNum, column=17).value = self.cirConEl_chUpper[i]
                worksheet.cell(row=rowNum, column=18).value = self.gen_convex_distr_cum[i]
                # Elongation
                worksheet.cell(row=rowNum, column=20).value = i + 1
                worksheet.cell(row=rowNum, column=21).value = self.cirConEl_chLower[i]
                worksheet.cell(row=rowNum, column=22).value = self.cirConEl_chCentre[i]
                worksheet.cell(row=rowNum, column=23).value = self.cirConEl_chUpper[i]
                worksheet.cell(row=rowNum, column=24).value = self.gen_elong_distr_cum[i]
                # Solidity
                worksheet.cell(row=rowNum, column=26).value = i + 1
                worksheet.cell(row=rowNum, column=27).value = self.cirConEl_chLower[i]
                worksheet.cell(row=rowNum, column=28).value = self.cirConEl_chCentre[i]
                worksheet.cell(row=rowNum, column=29).value = self.cirConEl_chUpper[i]
                worksheet.cell(row=rowNum, column=30).value = self.gen_solid_distr_cum[i]             
            rowNum += 1

        xlsxFilename = self.fileName[0:-4] + '.xlsx'
        workbook.save(xlsxFilename)
    
    def timer_event(self):
        self.elapsedTime += 1
        timeString = self.make_label_for_time(self.elapsedTime)
        self.edt_elapsedTime.setText(timeString)

    def make_label_for_time(self, value):
        """Method for creating good-loking string with time"""
        hours = value // 3600
        value = value - hours * 3600
        minutes = value // 60
        value = value - minutes * 60
        seconds = value
        return "{0:d}:{1:02d}:{2:02d}".format(hours, minutes, seconds)

    def show_error_window(self, text):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText(text)
        msg.setWindowTitle("Error")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()

    def show_information_window(self, text):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(text)
        msg.setWindowTitle("Information")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()

    def show_help(self):
        path = os.getcwd() + '\Help\Help_ParticlesGenerator.pdf'
        os.startfile(path)
        
    def show_about(self):
        self.about = About()
   
    def center_window(self):
        """Method for center the main window on the desktop"""
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

if __name__ == '__main__':
    def run_app():
        app = QApplication(sys.argv)
        app.setFont(QFont('Arial', 11)) 
        mainWindow = ParticlesGenerator()
        app.exec_()
    run_app()
    #sys.exit(app.exec_())