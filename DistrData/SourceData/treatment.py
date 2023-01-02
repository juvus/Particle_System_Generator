"""Program for treatment the raw Grafula data and for producing the PSD data suitable
   for the following usage (partigle generator tool)"""

import numpy as np
import xlrd
from openpyxl import load_workbook
from os.path import join, dirname, abspath
from gekko import GEKKO
import math as m
import matplotlib.pyplot as plt

# Class of the particle system
class Treatment():
    def __init__(self):
        self.excelFilename = 'TimeArray.xlsx'  # Filename of the Excel file with data
        self.inputStartRow = 7  # Input array start row (from 0)
        self.inputEndRow = None  # Input array end row (edit HERE)
        self.outputStartRow = 7  # Output array start row (from 0)
        self.outputEndRow = None  # Output array end row
        # Data arrays:
        self.inputX = []  # Input X data
        self.inputY = []  # Input Y data
        self.outputX = []  # Output X data
        self.outputY = []  # Output Y data
        # Critical points (0 or 100):
        self.last_0_i = 0
        self.last_0_X = 0
        self.last_0_Y = 0
        self.first_100_X = 0
        self.first_100_Y = 0
        self.first_100_i = 0

    def read_excel_data(self):
        """Method for reading the data form the Excel"""
        # Determine the excel filename
        fname = join(dirname(abspath(__file__)), self.excelFilename)
        workbook = xlrd.open_workbook(fname)
        sheetNames = workbook.sheet_names()
        xlSheet = workbook.sheet_by_name(sheetNames[0])
        
        # Determine inputEndRow and outputEndRow
        self.inputEndRow = self.inputStartRow
        emptyFlag = False
        while not emptyFlag:
            try:
                temp = xlSheet.cell(self.inputEndRow, 1).value
                if temp == "":
                    emptyFlag = True
                else:
                    self.inputEndRow += 1
            except:
                emptyFlag = True
        
        self.outputEndRow = self.outputStartRow
        emptyFlag = False
        while not emptyFlag:
            try:
                temp = xlSheet.cell(self.outputEndRow, 4).value
                if temp == "":
                    emptyFlag = True
                else:
                    self.outputEndRow += 1
            except:
                emptyFlag = True
                
        # Grab the input x, y data:
        for i in range(self.inputStartRow, self.inputEndRow):
            self.inputX.append(float(xlSheet.cell(i, 1).value))
            self.inputY.append(float(xlSheet.cell(i, 2).value))
        # Grab the output x data:
        for i in range(self.outputStartRow, self.outputEndRow):
            self.outputX.append(float(xlSheet.cell(i, 4).value))

    def make_treatment(self):
        """Method for the main treatment of the input data"""

        # Determine X and Y of the last Y=0 element in PSD
        for i in range(0, len(self.inputY)):
            if self.inputY[i] != 0:
                if i == 0:
                    self.last_0_i = 0
                    self.last_0_X = self.inputX[0]
                    self.last_0_Y = self.inputY[0]
                    break
                else:
                    self.last_0_i = i - 1
                    self.last_0_X = self.inputX[i - 1]
                    self.last_0_Y = self.inputY[i - 1]
                    break
        # Determine X and Y of the first Y=100 element in PSD
        for i in range(0, len(self.inputY)):
            if self.inputY[i] == 100:
                self.first_100_i = i
                self.first_100_X = self.inputX[i]
                self.first_100_Y = self.inputY[i]
                break

        # Create cubic spline
        gk = GEKKO()
        gk.x = gk.Param(value=np.asarray(self.outputX))
        gk.y = gk.Var()
        gk.cspline(gk.x, gk.y, np.asarray(self.inputX), np.asarray(self.inputY))
        gk.options.IMODE = 2
        gk.solve(disp=False)

        # Post treatment
        for i in range(0, len(self.outputX)):
            if (self.outputX[i] < self.last_0_X):
                self.outputY.append(0)
            elif (self.outputX[i] > self.first_100_X):
                self.outputY.append(100)
            else:
                if gk.y[i] < 0:
                    self.outputY.append(0)
                elif gk.y[i] > 100:
                    self.outputY.append(100)
                else:
                    self.outputY.append(gk.y[i])

    def write_excel_data(self):
        """Method for writing the data to the Excel"""
        fname = join(dirname(abspath(__file__)), self.excelFilename)
        workbook = load_workbook(filename=fname)
        worksheet = workbook["Data"]
        for i in range(0, len(self.outputY)):
            rowNum = self.outputStartRow + i + 1
            worksheet.cell(row=rowNum, column=6).value = self.outputY[i]
        workbook.save(fname)

    def make_plots(self):
        plt.figure(1)
        plt.plot(self.inputX, self.inputY, label='Input data', color='r')
        plt.plot(self.outputX, self.outputY, label='Output data', color='b')
        plt.xlabel('x parameter')
        plt.ylabel('y parameter')
        plt.legend()
        
        if max(self.inputX) <= 1: # Circularity, Convexity, Elongation
            plt.xlim([0, 1])
        elif max(self.inputX) > 1: # CE Diameter
            plt.xscale('log')
            plt.xlim([0.01, 10000])
        plt.show()

if __name__ == "__main__":
    treatment = Treatment()
    treatment.read_excel_data()
    treatment.make_treatment()
    treatment.write_excel_data()
    treatment.make_plots()







